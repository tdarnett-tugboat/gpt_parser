import argparse
import concurrent
from concurrent.futures import ThreadPoolExecutor

import openai
from openpyxl import load_workbook

# Set a prompt token to identify the cells that contain prompts
PROMPT_TOKEN = "{OT_GPT}"


def main(
    input_file: str, sheet: str, output_file: str, api_key: str, prompt_token: str
):
    workbook = load_workbook(input_file, data_only=True)
    worksheet = workbook[sheet]

    # extract prompts from the spreadsheet
    prompts_by_cell_coordinate = _extract_prompts(worksheet, prompt_token)

    full_message = f"Identified {len(prompts_by_cell_coordinate)} prompts in the spreadsheet. Continue?\n"
    if input(full_message) != "yes":
        raise ValueError("Operation cancelled.")

    # configure openai client
    openai.api_key = api_key

    # Resolve prompts and map the answers to the cells of the original prompt
    answers_by_cell_coordinate = _answers_by_cell_coordinate(prompts_by_cell_coordinate)

    # write the answers to the corresponding cells in the worksheet
    for cell, answer in answers_by_cell_coordinate.items():
        worksheet[cell] = answer

    # save workbook to the specified output file
    workbook.save(output_file)


def _answers_by_cell_coordinate(prompts_by_cell_coordinate):
    resolved_prompts_by_cell_coordinate = {}
    # resolve prompts concurrently
    with ThreadPoolExecutor() as executor:
        futures = {}
        for cell, prompt in prompts_by_cell_coordinate.items():
            futures[executor.submit(_resolve_prompt, prompt)] = cell

        for future in concurrent.futures.as_completed(futures):
            cell = futures[future]
            resolved_prompts_by_cell_coordinate[cell] = future.result()

    return resolved_prompts_by_cell_coordinate


def _resolve_prompt(prompt: str) -> str:
    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "user", "content": prompt},
        ],
        temperature=0.8,
        max_tokens=60,
        top_p=1.0,
        frequency_penalty=0.0,
        presence_penalty=0.0,
    )
    answer = response["choices"][0]["message"]["content"]
    return answer


def _extract_prompts(worksheet, token):
    prompts_by_cell_coordinate = {}
    for row in worksheet.iter_rows():
        for c in row:
            if c.value and token in c.value:
                prompts_by_cell_coordinate[c.coordinate] = c.value.replace(token, "")
    return prompts_by_cell_coordinate


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Replace spreadsheet cells with ChatGPT responses"
    )
    parser.add_argument(
        "--input_file", help="file path of the input spreadsheet", required=True
    )
    parser.add_argument("--sheet", help="Sheet to read", required=True)
    parser.add_argument(
        "--output_file",
        help="file path of where to export the new spreadsheet",
        required=True,
    )
    parser.add_argument("--api_key", help="OpenAI API Key", required=True)
    parser.add_argument(
        "--prompt_token",
        help=f"Token for signalling prompt cells. Default '{PROMPT_TOKEN}'",
        required=False,
    )

    args = parser.parse_args()

    prompt_token = args.prompt_token or PROMPT_TOKEN
    main(args.input_file, args.sheet, args.output_file, args.api_key, prompt_token)
